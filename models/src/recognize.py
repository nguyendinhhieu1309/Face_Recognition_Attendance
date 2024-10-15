import cv2
import numpy as np
from mtcnn import MTCNN
from keras_facenet import FaceNet
from sklearn.preprocessing import Normalizer
import pickle
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pandas as pd
import os
from PIL import ImageFont, ImageDraw, Image

def recognize(student_data):
    url = 'http://192.168.0.100:8080/video'
    vid = cv2.VideoCapture(url)

    if not vid.isOpened():
        print(f"Error opening video stream or file with URL: {url}")
        return

    detector = MTCNN()
    embedder = FaceNet()
    in_encoder = Normalizer(norm='l2')
    
    # Load model SVM và encoder
    filename_path = r'F:\SUMMER24\CPV301\Face_Recognition_Attendance_System/models/weights/svm_saved.sav'
    loaded_model = pickle.load(open(filename_path, 'rb'))
    encoder_path = r'F:\SUMMER24\CPV301\Face_Recognition_Attendance_System/models/weights/out_encoder.pkl'
    out_encoder = pickle.load(open(encoder_path, 'rb'))
    
    # Load dữ liệu sinh viên từ file CSV
    df_students = pd.read_csv(r'F:\SUMMER24\CPV301\Face_Recognition_Attendance_System/students.csv')
    print(f"DataFrame columns: {df_students.columns}")
    print(f"DataFrame head:\n{df_students.head()}")

    # Kiểm tra xem file attendance.xlsx đã tồn tại chưa
    attendance_file = 'F:\SUMMER24\CPV301\Face_Recognition_Attendance_System/attendance.xlsx'
    if os.path.exists(attendance_file):
        wb = load_workbook(attendance_file)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = 'Student ID'
        sheet['B1'] = 'Name'
        sheet['C1'] = 'Time'

    attended = set()

    while True:
        ret, frame = vid.read()
        if not ret:
            print("Failed to capture image")
            break

        image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        result = detector.detect_faces(image)

        if not result:
            print("No faces detected.")
        
        for each_result in result:
            bounding_box = each_result['box']
            confidence = each_result['confidence']
            
            # In ra thông tin bounding box và confidence để kiểm tra
            print(f"Bounding box: {bounding_box}, Confidence: {confidence}")

            if confidence > 0.90:  # Kiểm tra ngưỡng độ tin cậy
                crop_img = image[bounding_box[1]:bounding_box[1] + bounding_box[3], bounding_box[0]:bounding_box[0] + bounding_box[2]]
                resized_image = cv2.resize(crop_img, (160, 160))
                embeddings = embedder.embeddings([resized_image])
                testX = in_encoder.transform(embeddings)
                y_pred_encoded = loaded_model.predict(testX)
                y_pred = out_encoder.inverse_transform(y_pred_encoded)[0]  # Convert back to original label

                print(f"Predicted Student ID: {y_pred}")

                # Lấy thông tin sinh viên từ df_students dựa trên y_pred (mã sinh viên dự đoán)
                student_info = df_students[df_students['Student ID'] == y_pred]
                print(f"Student Info for ID {y_pred}: {student_info}")

                if not student_info.empty:
                    student_name = student_info.iloc[0]['Name']
                    print(f"Recognized Student Name: {student_name}")

                    time_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    # Vẽ hình chữ nhật và nhãn với font Unicode
                    font_path = "F:\SUMMER24\CPV301\Face_Recognition_Attendance_System\models\src\Arial.ttf"  # Đường dẫn tới tệp font Arial
                    font = ImageFont.truetype(font_path, 24)
                    pil_image = Image.fromarray(frame)
                    draw = ImageDraw.Draw(pil_image)
                    
                    draw.rectangle([(bounding_box[0], bounding_box[1]), (bounding_box[0] + bounding_box[2], bounding_box[1] + bounding_box[3])], outline="blue", width=2)
                    draw.text((bounding_box[0], bounding_box[1] - 30), f'{y_pred} - {student_name}', font=font, fill=(0, 255, 0, 255))
                    frame = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)
                    
                    if y_pred not in attended:
                        attended.add(y_pred)
                        sheet.append([y_pred, student_name, time_now])
                        print(f"Added to attendance: ID = {y_pred}, Name = {student_name}, Time = {time_now}")
                else:
                    print(f"Student ID {y_pred} not found in database.")
                    print("Current student data:")
                    print(df_students)
            else:
                print(f"Low confidence: {confidence}")

        cv2.namedWindow('frame', cv2.WINDOW_NORMAL)
        cv2.resizeWindow('frame', 800, 600)
        cv2.imshow('frame', cv2.cvtColor(frame, cv2.COLOR_RGB2BGR))

        if cv2.waitKey(1) & 0xFF == ord('q'):
            wb.save(filename=attendance_file)
            print("Attendance saved.")
            break

    vid.release()
    cv2.destroyAllWindows()
    




